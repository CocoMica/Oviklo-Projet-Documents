import openpyxl

def Search_For_A_record(path):

    try:
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        for r in range(1,sheet_obj.max_row):
            #r_new = r+1
            temp_PO = str(sheet_obj.cell(row=r, column=10).value)
            if temp_PO == '4501297680':
                print("found an entry")
    except:
        print("error")

def Open_WB(path):
    wb_obj = openpyxl.load_workbook(path)

