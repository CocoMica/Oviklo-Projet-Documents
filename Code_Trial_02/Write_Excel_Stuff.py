from openpyxl import *
from datetime import datetime
global sheet
global workbook
global file_name
global Excel_Path
Excel_Name = "Oviklo_Inventory_Records"


def get_Excel_Path(path):
    global Excel_Path
    Excel_Path = path



def get_month(name):
    mth = str(datetime.today().month)
    week = str(datetime.today().isocalendar()[1])
    if len(mth) < 2:
        mth = "0"+mth
    #total_name = Excel_Path + str(name) + "-" + str(datetime.today().year) + "-" + mth+".xlsx" #if needed by month
    total_name = Excel_Path + str(name) + "-" + str(datetime.today().year)+".xlsx" #if needed by year
    return total_name


def create_workbook(name):
    global sheet
    global workbook
    global file_name
    file_name = get_month(name)
    try:
        workbook = load_workbook(filename=file_name)
        sheet = workbook.active
        sheet.protection.password = "12345"
        sheet["A1"] = "Date Stamp"
        sheet["B1"] = "Time Stamp"
        sheet["C1"] = "Label issue number" #to track and cross reference number of labels created for a specific label.
        sheet["D1"] = "EPF"
        sheet["E1"] = "Barcode"
        sheet["F1"] = "Quantity"
        sheet["G1"] = "Box Quantity"
        workbook.save(filename=file_name)

    except:
        workbook = Workbook()
        sheet = workbook.active
        sheet.protection.password = "12345"
        sheet["A1"] = "Date Stamp"
        sheet["B1"] = "Time Stamp"
        sheet["C1"] = "Label issue number"
        sheet["D1"] = "EPF"
        sheet["E1"] = "Barcode"
        sheet["F1"] = "Quantity"
        sheet["G1"] = "Box Quantity"
        workbook.save(filename=file_name)


def gather_data():
    global sheet
    global workbook
    create_workbook(Excel_Name)
    global file_name
    list_DateStamp = []
    list_TimeStamp = []
    list_PO = []
    list_CN = []
    list_UQ = []
    list_SN = []
    list_WT = []
    current_Row = 0
    for row in range(1, sheet.max_row + 1):
        current_Row += 1
        DateStamp_Cell_Name = "A"+str(current_Row)
        TimeStamp_Cell_Name = "B"+str(current_Row)
        PO_Cell_Name = "C"+str(current_Row)
        CN_Cell_Name = "D"+str(current_Row)
        UQ_Cell_Name = "E"+str(current_Row)
        SN_Cell_Name = "F"+str(current_Row)
        WT_Cell_Name = "G"+str(current_Row)
        list_DateStamp.append(sheet[DateStamp_Cell_Name].value)
        list_TimeStamp.append(sheet[TimeStamp_Cell_Name].value)
        list_PO.append(sheet[PO_Cell_Name].value)
        list_CN.append(sheet[CN_Cell_Name].value)
        list_UQ.append(sheet[UQ_Cell_Name].value)
        list_SN.append(sheet[SN_Cell_Name].value)
        list_WT.append(sheet[WT_Cell_Name].value)

    workbook.close()
    return current_Row, list_DateStamp, list_TimeStamp, list_PO, list_CN, list_UQ, list_SN, list_WT


def write_to_last_column_workbook2(Po_Num, Carton_Num, Unit_Qty, Style_Num, Weight):
    try:
        global sheet
        global workbook
        global file_name
        output_state = "0"
        create_workbook(Excel_Name)
        Both_Entries_Not_Already_Recorded = True
        for r in range(1, sheet.max_row + 1):
            Temp_PO = sheet.cell(row=r, column=3)
            Temp_CN = sheet.cell(row=r, column=4)
            if Temp_PO.value == Po_Num and Temp_CN.value == Carton_Num:
                Both_Entries_Not_Already_Recorded = False

        if (Both_Entries_Not_Already_Recorded):
            Ds_Cell = "A"+str(sheet.max_row + 1)
            Ts_Cell = "B"+str(sheet.max_row + 1)
            Po_Cell = "C"+str(sheet.max_row + 1)
            Cn_Cell = "D"+str(sheet.max_row + 1)
            Qty_Cell = "E"+str(sheet.max_row + 1)
            Style_Cell = "F"+str(sheet.max_row + 1)
            W_Cell = "G"+str(sheet.max_row + 1)
            now = datetime.now()
            dt_string = now.strftime("%d/%m/%Y")
            tm_string = now.strftime("%H:%M:%S")

            sheet[Ds_Cell] = dt_string
            sheet[Ts_Cell] = tm_string
            sheet[Po_Cell] = Po_Num
            sheet[Cn_Cell] = Carton_Num
            sheet[Qty_Cell] = Unit_Qty
            sheet[Style_Cell] = Style_Num
            sheet[W_Cell] = Weight
            workbook.save(filename=file_name)
            workbook.close()
            output_state = "0"
            return output_state, None
        else:
            output_state = "1"
            return output_state, None
    except Exception as Err:
        print("Error Writing data to Excel: ", Err)
        output_state = "2"
        return output_state, repr(Err)

def Create_Excel_Document(Path, Name):
    get_Excel_Path(Path)
    try:
        create_workbook(Name)
        print("Excel file created at: ", Excel_Path)
        return True
    except:
        return False